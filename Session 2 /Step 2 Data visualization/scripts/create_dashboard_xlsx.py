#!/usr/bin/env python3
"""
Create a WHU-styled Excel dashboard from a JSON spec.

Usage:
  python scripts/create_dashboard_xlsx.py --spec assets/example_xlsx_spec.json --out out.xlsx
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WHU_DARK_BLUE = "054696"
WHU_LIGHT_BLUE = "00A5DC"
WHU_LIGHT_GREY = "E6EBF0"
BLACK = "000000"
WHITE = "FFFFFF"

def build(spec_path: Path, out_path: Path):
    spec = json.loads(spec_path.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    data_ws = wb.create_sheet("Data")

    # Title
    ws["A1"] = spec.get("title", "Dashboard")
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color=WHU_DARK_BLUE)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # KPI cards (simple grid)
    kpis = spec.get("kpis", [])
    start_row = 3
    col = 1
    thin = Side(style="thin", color=WHU_LIGHT_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, k in enumerate(kpis[:6]):
        r = start_row + (i//3)*4
        c = col + (i%3)*3
        label_cell = ws.cell(row=r, column=c, value=str(k.get("label","")))
        value_cell = ws.cell(row=r+1, column=c, value=str(k.get("value","")))
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+2)
        ws.merge_cells(start_row=r+1, start_column=c, end_row=r+2, end_column=c+2)

        label_cell.font = Font(name="Arial", size=11, bold=True, color=WHU_DARK_BLUE)
        value_cell.font = Font(name="Arial", size=18, bold=True, color=WHU_LIGHT_BLUE)
        label_cell.fill = PatternFill("solid", fgColor=WHU_LIGHT_GREY)
        value_cell.fill = PatternFill("solid", fgColor=WHITE)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        # apply border to merged region
        for rr in range(r, r+3):
            for cc in range(c, c+3):
                ws.cell(row=rr, column=cc).border = border

    # Data table
    data = spec.get("data", {})
    cols = data.get("columns", [])
    rows = data.get("rows", [])

    for j, colname in enumerate(cols, start=1):
        cell = data_ws.cell(row=1, column=j, value=str(colname))
        cell.font = Font(name="Arial", bold=True, color=WHU_DARK_BLUE)
        cell.fill = PatternFill("solid", fgColor=WHU_LIGHT_GREY)
        cell.alignment = Alignment(horizontal="left")

    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            cell = data_ws.cell(row=i, column=j, value=val)
            cell.font = Font(name="Arial", color=BLACK)
            cell.alignment = Alignment(horizontal="left")

    # Column widths
    for ws_ in [ws, data_ws]:
        for col_idx in range(1, 9):
            ws_.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(str(out_path))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--spec", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    args = ap.parse_args()
    build(args.spec, args.out)

if __name__ == "__main__":
    main()
